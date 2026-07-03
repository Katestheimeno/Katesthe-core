"""Tests for CSV/XLSX export helpers and formula-injection sanitization.
Path: utils/tests/test_export.py
"""

from io import BytesIO

from openpyxl import load_workbook

from utils.export import (
    _is_number,
    _sanitize_cell,
    _sanitize_filename,
    csv_response,
    xlsx_response,
)


class TestIsNumber:
    """`_is_number` recognizes numeric literals only."""

    def test_positive_integer_string_is_number(self):
        assert _is_number("5") is True

    def test_negative_integer_string_is_number(self):
        assert _is_number("-5") is True

    def test_non_numeric_string_is_not_number(self):
        assert _is_number("-abc") is False

    def test_none_is_not_number(self):
        assert _is_number(None) is False


class TestSanitizeCell:
    """`_sanitize_cell` neutralizes formula-injection payloads."""

    def test_equals_prefixed_value_is_quote_prefixed(self):
        assert _sanitize_cell("=x") == "'=x"

    def test_at_prefixed_value_is_quote_prefixed(self):
        assert _sanitize_cell("@x") == "'@x"

    def test_plus_prefixed_value_is_quote_prefixed(self):
        assert _sanitize_cell("+x") == "'+x"

    def test_tab_prefixed_value_is_quote_prefixed(self):
        assert _sanitize_cell("\tx") == "'\tx"

    def test_carriage_return_prefixed_value_is_quote_prefixed(self):
        assert _sanitize_cell("\rx") == "'\rx"

    def test_negative_number_is_not_prefixed(self):
        assert _sanitize_cell("-5") == "-5"

    def test_negative_non_numeric_value_is_quote_prefixed(self):
        assert _sanitize_cell("-abc") == "'-abc"

    def test_none_value_becomes_empty_string(self):
        assert _sanitize_cell(None) == ""

    def test_plain_value_is_unchanged(self):
        assert _sanitize_cell("hello") == "hello"


class TestSanitizeFilename:
    """`_sanitize_filename` strips unsafe path/whitespace characters."""

    def test_path_traversal_and_spaces_are_stripped(self):
        result = _sanitize_filename("../../etc/pa ss")
        assert "/" not in result
        assert " " not in result

    def test_empty_result_falls_back_to_export(self):
        assert _sanitize_filename("   ///") == "export"

    def test_safe_name_is_unchanged(self):
        assert _sanitize_filename("report_2026-07-02") == "report_2026-07-02"

    def test_dotted_path_traversal_collapses_repeated_dots(self):
        result = _sanitize_filename("../../etc/passwd")
        assert ".." not in result
        assert "/" not in result


class TestCsvResponse:
    """`csv_response` produces a sanitized CSV attachment."""

    def test_content_type_is_csv(self):
        resp = csv_response("report", ["a", "b"], [{"a": "=SUM(1)", "b": "x"}])
        assert resp["Content-Type"].startswith("text/csv")

    def test_formula_cell_is_quote_prefixed_in_body(self):
        resp = csv_response("report", ["a", "b"], [{"a": "=SUM(1)", "b": "x"}])
        assert "'=SUM(1)" in resp.content.decode()

    def test_content_disposition_has_sanitized_filename(self):
        resp = csv_response("report", ["a", "b"], [{"a": "=SUM(1)", "b": "x"}])
        assert resp["Content-Disposition"] == 'attachment; filename="report.csv"'

    def test_list_rows_are_written_positionally(self):
        resp = csv_response("report", ["a", "b"], [["1", "2"]])
        body = resp.content.decode()
        assert "1,2" in body

    def test_dangerous_filename_is_sanitized_in_disposition(self):
        resp = csv_response("../evil name", ["a"], [{"a": "1"}])
        disposition = resp["Content-Disposition"]
        assert "/" not in disposition
        assert " name" not in disposition


class TestXlsxResponse:
    """`xlsx_response` produces a sanitized XLSX attachment."""

    def test_content_type_is_xlsx_mime(self):
        resp = xlsx_response("data", ["a"], [{"a": "=1+1"}])
        assert resp["Content-Type"] == (
            "application/vnd.openxmlformats-officedocument"
            ".spreadsheetml.sheet"
        )

    def test_body_is_non_empty(self):
        resp = xlsx_response("data", ["a"], [{"a": "=1+1"}])
        assert len(resp.content) > 0

    def test_injected_cell_is_quote_prefixed_when_reloaded(self):
        resp = xlsx_response("data", ["a"], [{"a": "=1+1"}])
        wb = load_workbook(BytesIO(resp.content))
        ws = wb.active
        assert ws["A2"].value == "'=1+1"

    def test_content_disposition_has_xlsx_extension(self):
        resp = xlsx_response("data", ["a"], [{"a": "1"}])
        assert resp["Content-Disposition"] == 'attachment; filename="data.xlsx"'
